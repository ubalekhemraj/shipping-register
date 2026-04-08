"""
Medical Shop Shipping Register
Full-stack Reflex application
"""

import reflex as rx
from datetime import datetime, date
from typing import Optional
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import base64


# ─────────────────────────────────────────
# DATABASE MODEL
# ─────────────────────────────────────────

class ShippingEntry(rx.Model, table=True):
    """Represents one shipping entry in the register."""

    name: str
    entry_datetime: str          # ISO string "YYYY-MM-DD HH:MM"
    from_location: str
    to_location: str
    shipping_address: str
    cartons_count: int
    shipping_company: str
    created_at: str


# ─────────────────────────────────────────
# STATE
# ─────────────────────────────────────────

class State(rx.State):
    # ── form fields ──────────────────────
    f_name: str = ""
    f_entry_datetime: str = ""
    f_from_location: str = ""
    f_to_location: str = ""
    f_shipping_address: str = ""
    f_cartons_count: str = "1"
    f_shipping_company: str = ""

    # ── UI flags ─────────────────────────
    show_form: bool = False
    form_error: str = ""
    toast_msg: str = ""
    toast_type: str = ""   # "success" | "error"

    # ── filter ───────────────────────────
    filter_date_from: str = str(date.today())
    filter_date_to: str = str(date.today())

    # ── data ─────────────────────────────
    entries: list[dict] = []

    # ── summary ──────────────────────────
    summary_total: int = 0
    summary_cartons: int = 0
    summary_companies: str = ""

    # ─────────────────────────────────────
    def on_load(self):
        self.load_entries()

    def toggle_form(self):
        self.show_form = not self.show_form
        self.form_error = ""
        if self.show_form:
            # Pre-fill datetime to now
            self.f_entry_datetime = datetime.now().strftime("%Y-%m-%dT%H:%M")

    def set_f_name(self, v: str):
        self.f_name = v

    def set_f_entry_datetime(self, v: str):
        self.f_entry_datetime = v

    def set_f_from_location(self, v: str):
        self.f_from_location = v

    def set_f_to_location(self, v: str):
        self.f_to_location = v

    def set_f_shipping_address(self, v: str):
        self.f_shipping_address = v

    def set_f_cartons_count(self, v: str):
        self.f_cartons_count = v

    def set_f_shipping_company(self, v: str):
        self.f_shipping_company = v

    def set_filter_date_from(self, v: str):
        self.filter_date_from = v

    def set_filter_date_to(self, v: str):
        self.filter_date_to = v

    def apply_filter(self):
        self.load_entries()

    def reset_to_today(self):
        self.filter_date_from = str(date.today())
        self.filter_date_to = str(date.today())
        self.load_entries()

    # ─────────────────────────────────────
    def load_entries(self):
        with rx.session() as session:
            all_entries = session.exec(
                ShippingEntry.select().order_by(ShippingEntry.entry_datetime.desc())
            ).all()

            result = []
            for e in all_entries:
                # entry_datetime stored as "YYYY-MM-DD HH:MM"
                entry_date = e.entry_datetime[:10]
                if self.filter_date_from <= entry_date <= self.filter_date_to:
                    result.append(
                        {
                            "id": e.id,
                            "name": e.name,
                            "entry_datetime": e.entry_datetime,
                            "from_location": e.from_location,
                            "to_location": e.to_location,
                            "shipping_address": e.shipping_address,
                            "cartons_count": e.cartons_count,
                            "shipping_company": e.shipping_company,
                        }
                    )

            self.entries = result
            self._refresh_summary()

    def _refresh_summary(self):
        self.summary_total = len(self.entries)
        self.summary_cartons = sum(e["cartons_count"] for e in self.entries)
        companies = list({e["shipping_company"] for e in self.entries if e["shipping_company"]})
        self.summary_companies = ", ".join(companies) if companies else "—"

    # ─────────────────────────────────────
    def submit_entry(self):
        # Validate required fields
        missing = []
        if not self.f_name.strip():
            missing.append("Name")
        if not self.f_from_location.strip():
            missing.append("From Location")
        if not self.f_to_location.strip():
            missing.append("To Location")
        if not self.f_shipping_address.strip():
            missing.append("Shipping Address")
        if not self.f_shipping_company.strip():
            missing.append("Shipping Company")

        if missing:
            self.form_error = "Required: " + ", ".join(missing)
            return

        try:
            cartons = int(self.f_cartons_count) if self.f_cartons_count else 0
        except ValueError:
            self.form_error = "Cartons must be a number"
            return

        # Parse datetime
        if self.f_entry_datetime:
            try:
                dt = datetime.strptime(self.f_entry_datetime, "%Y-%m-%dT%H:%M")
                dt_str = dt.strftime("%Y-%m-%d %H:%M")
            except Exception:
                dt_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        else:
            dt_str = datetime.now().strftime("%Y-%m-%d %H:%M")

        with rx.session() as session:
            entry = ShippingEntry(
                name=self.f_name.strip(),
                entry_datetime=dt_str,
                from_location=self.f_from_location.strip(),
                to_location=self.f_to_location.strip(),
                shipping_address=self.f_shipping_address.strip(),
                cartons_count=cartons,
                shipping_company=self.f_shipping_company.strip(),
                created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            )
            session.add(entry)
            session.commit()

        # Reset form
        self.f_name = ""
        self.f_entry_datetime = datetime.now().strftime("%Y-%m-%dT%H:%M")
        self.f_from_location = ""
        self.f_to_location = ""
        self.f_shipping_address = ""
        self.f_cartons_count = "1"
        self.f_shipping_company = ""
        self.form_error = ""
        self.show_form = False

        self.load_entries()

    # ─────────────────────────────────────
    def delete_entry(self, entry_id: int):
        with rx.session() as session:
            entry = session.get(ShippingEntry, entry_id)
            if entry:
                session.delete(entry)
                session.commit()
        self.load_entries()

    # ─────────────────────────────────────
    def export_excel(self):
        if not self.entries:
            return

        wb = openpyxl.Workbook()

        # ── Sheet 1: Register ─────────────────────────────────────
        ws = wb.active
        ws.title = "Shipping Register"

        # Header styling
        header_fill = PatternFill("solid", fgColor="1E40AF")   # deep blue
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="BFDBFE")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            "S.No", "Name", "Date & Time", "From Location",
            "To Location", "Shipping Address", "Cartons", "Shipping Company",
        ]
        col_widths = [6, 22, 18, 22, 22, 30, 10, 22]

        for col_num, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[get_column_letter(col_num)].width = width

        ws.row_dimensions[1].height = 30

        # Data rows
        row_fill_even = PatternFill("solid", fgColor="EFF6FF")
        row_fill_odd = PatternFill("solid", fgColor="FFFFFF")
        data_font = Font(name="Calibri", size=10)
        data_align = Alignment(vertical="center", wrap_text=True)

        for idx, entry in enumerate(self.entries, start=1):
            row = idx + 1
            fill = row_fill_even if idx % 2 == 0 else row_fill_odd
            row_data = [
                idx,
                entry["name"],
                entry["entry_datetime"],
                entry["from_location"],
                entry["to_location"],
                entry["shipping_address"],
                entry["cartons_count"],
                entry["shipping_company"],
            ]
            for col_num, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col_num, value=value)
                cell.font = data_font
                cell.fill = fill
                cell.alignment = data_align
                cell.border = border
            ws.row_dimensions[row].height = 20

        # Freeze header row
        ws.freeze_panes = "A2"

        # ── Sheet 2: Summary ─────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        title_font = Font(name="Calibri", bold=True, size=14, color="1E3A8A")
        label_font = Font(name="Calibri", bold=True, size=11)
        value_font = Font(name="Calibri", size=11)
        section_fill = PatternFill("solid", fgColor="DBEAFE")

        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 35

        ws2.merge_cells("A1:B1")
        ws2["A1"] = "🏥 Medical Shop — Shipping Register Summary"
        ws2["A1"].font = title_font
        ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws2["A1"].fill = PatternFill("solid", fgColor="BFDBFE")
        ws2.row_dimensions[1].height = 32

        summary_rows = [
            ("Date Range", f"{self.filter_date_from}  →  {self.filter_date_to}"),
            ("Total Entries", str(self.summary_total)),
            ("Total Cartons / Boxes", str(self.summary_cartons)),
            ("Shipping Companies", self.summary_companies),
            ("Report Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

        for row_idx, (label, value) in enumerate(summary_rows, start=3):
            lc = ws2.cell(row=row_idx, column=1, value=label)
            vc = ws2.cell(row=row_idx, column=2, value=value)
            lc.font = label_font
            vc.font = value_font
            lc.fill = section_fill
            lc.alignment = Alignment(vertical="center")
            vc.alignment = Alignment(vertical="center")
            ws2.row_dimensions[row_idx].height = 22

        # ── Company-wise breakdown ─────────────────────────────
        ws2.cell(row=9, column=1, value="Company-wise Breakdown").font = Font(
            name="Calibri", bold=True, size=11, color="1E40AF"
        )

        company_data: dict[str, dict] = {}
        for e in self.entries:
            co = e["shipping_company"] or "Unknown"
            if co not in company_data:
                company_data[co] = {"count": 0, "cartons": 0}
            company_data[co]["count"] += 1
            company_data[co]["cartons"] += e["cartons_count"]

        ch_row = 10
        ws2.cell(row=ch_row, column=1, value="Company").font = Font(bold=True, size=10, color="FFFFFF")
        ws2.cell(row=ch_row, column=1).fill = PatternFill("solid", fgColor="1E40AF")
        ws2.cell(row=ch_row, column=2, value="Entries | Cartons").font = Font(bold=True, size=10, color="FFFFFF")
        ws2.cell(row=ch_row, column=2).fill = PatternFill("solid", fgColor="1E40AF")

        for i, (co, stats) in enumerate(company_data.items(), start=ch_row + 1):
            ws2.cell(row=i, column=1, value=co).font = Font(size=10)
            ws2.cell(row=i, column=2, value=f"{stats['count']} entries  |  {stats['cartons']} cartons").font = Font(size=10)

        # ── Serialize ─────────────────────────────────────────
        buf = BytesIO()
        wb.save(buf)
        excel_bytes = buf.getvalue()

        filename = f"shipping_register_{self.filter_date_from}_to_{self.filter_date_to}.xlsx"

        # Save to assets/exports so Reflex can serve it
        export_dir = os.path.join(os.getcwd(), "assets", "exports")
        os.makedirs(export_dir, exist_ok=True)
        filepath = os.path.join(export_dir, filename)
        with open(filepath, "wb") as f:
            f.write(excel_bytes)

        return rx.download(url=f"/exports/{filename}", filename=filename)


# ─────────────────────────────────────────
# STYLE CONSTANTS
# ─────────────────────────────────────────

BLUE = "#1E40AF"
BLUE_LIGHT = "#DBEAFE"
BLUE_MID = "#3B82F6"
GREEN = "#059669"
RED = "#DC2626"
GRAY = "#6B7280"
BG = "#F0F4FF"

CARD = dict(
    background="white",
    border_radius="12px",
    box_shadow="0 2px 8px rgba(30,64,175,0.08)",
    padding="24px",
    width="100%",
)

INPUT_STYLE = dict(
    border="1.5px solid #BFDBFE",
    border_radius="8px",
    padding="8px 12px",
    font_size="14px",
    width="100%",
    background="white",
    _focus={"outline": "none", "border_color": BLUE_MID, "box_shadow": "0 0 0 3px #DBEAFE"},
)

LABEL_STYLE = dict(font_size="13px", font_weight="600", color="#374151", margin_bottom="4px")

BTN_PRIMARY = dict(
    background=BLUE,
    color="white",
    border_radius="8px",
    padding="10px 20px",
    font_weight="600",
    font_size="14px",
    cursor="pointer",
    _hover={"background": "#1D3461"},
    border="none",
)

BTN_SUCCESS = dict(
    background=GREEN,
    color="white",
    border_radius="8px",
    padding="10px 20px",
    font_weight="600",
    font_size="14px",
    cursor="pointer",
    _hover={"background": "#047857"},
    border="none",
)

BTN_DANGER = dict(
    background="transparent",
    color=RED,
    border_radius="6px",
    padding="4px 10px",
    font_size="12px",
    font_weight="600",
    cursor="pointer",
    _hover={"background": "#FEE2E2"},
    border=f"1px solid {RED}",
)

BTN_OUTLINE = dict(
    background="white",
    color=BLUE,
    border_radius="8px",
    padding="10px 20px",
    font_weight="600",
    font_size="14px",
    cursor="pointer",
    _hover={"background": BLUE_LIGHT},
    border=f"1.5px solid {BLUE}",
)


# ─────────────────────────────────────────
# COMPONENTS
# ─────────────────────────────────────────

def labeled_input(label: str, placeholder: str, value, on_change, input_type: str = "text"):
    return rx.vstack(
        rx.text(label, **LABEL_STYLE),
        rx.input(
            placeholder=placeholder,
            value=value,
            on_change=on_change,
            type=input_type,
            **INPUT_STYLE,
        ),
        spacing="1",
        width="100%",
        align_items="flex-start",
    )


def summary_card(icon: str, label: str, value):
    return rx.box(
        rx.hstack(
            rx.text(icon, font_size="28px"),
            rx.vstack(
                rx.text(value, font_size="22px", font_weight="700", color=BLUE),
                rx.text(label, font_size="12px", color=GRAY, font_weight="500"),
                spacing="0",
                align_items="flex-start",
            ),
            spacing="3",
            align_items="center",
        ),
        background="white",
        border_radius="12px",
        border=f"1.5px solid {BLUE_LIGHT}",
        padding="16px 22px",
        min_width="160px",
        box_shadow="0 2px 6px rgba(30,64,175,0.07)",
    )


def entry_row(entry: dict):
    return rx.table.row(
        rx.table.cell(
            entry["name"],
            font_weight="600",
            color="#1E3A8A",
            padding="10px 14px",
        ),
        rx.table.cell(entry["entry_datetime"], color=GRAY, font_size="13px", padding="10px 14px"),
        rx.table.cell(
            rx.hstack(
                rx.text("↑", color=GREEN, font_weight="700"),
                rx.text(entry["from_location"], font_size="13px"),
                spacing="1",
            ),
            padding="10px 14px",
        ),
        rx.table.cell(
            rx.hstack(
                rx.text("↓", color=RED, font_weight="700"),
                rx.text(entry["to_location"], font_size="13px"),
                spacing="1",
            ),
            padding="10px 14px",
        ),
        rx.table.cell(entry["shipping_address"], font_size="12px", color=GRAY, padding="10px 14px", max_width="180px"),
        rx.table.cell(
            rx.badge(
                rx.text(f"{entry['cartons_count']} boxes"),
                color_scheme="blue",
                variant="soft",
            ),
            padding="10px 14px",
            text_align="center",
        ),
        rx.table.cell(
            rx.badge(entry["shipping_company"], color_scheme="green", variant="soft"),
            padding="10px 14px",
        ),
        rx.table.cell(
            rx.button(
                "✕ Delete",
                on_click=State.delete_entry(entry["id"]),
                **BTN_DANGER,
            ),
            padding="10px 14px",
        ),
        _hover={"background": "#F0F7FF"},
    )


def add_entry_form():
    return rx.cond(
        State.show_form,
        rx.box(
            rx.vstack(
                rx.hstack(
                    rx.text("📦 New Shipping Entry", font_size="18px", font_weight="700", color=BLUE),
                    rx.spacer(),
                    rx.button(
                        "✕",
                        on_click=State.toggle_form,
                        background="transparent",
                        border="none",
                        font_size="20px",
                        cursor="pointer",
                        color=GRAY,
                        _hover={"color": RED},
                    ),
                    width="100%",
                    align_items="center",
                ),
                rx.divider(border_color=BLUE_LIGHT),

                # Row 1
                rx.hstack(
                    labeled_input("Name *", "Consignment / person name", State.f_name, State.set_f_name),
                    labeled_input("Entry Date & Time *", "", State.f_entry_datetime, State.set_f_entry_datetime, "datetime-local"),
                    spacing="4",
                    width="100%",
                ),

                # Row 2
                rx.hstack(
                    labeled_input("From Location *", "Source city / warehouse", State.f_from_location, State.set_f_from_location),
                    labeled_input("To Location *", "Destination city / hospital", State.f_to_location, State.set_f_to_location),
                    spacing="4",
                    width="100%",
                ),

                # Row 3
                rx.hstack(
                    rx.vstack(
                        rx.text("Shipping Address *", **LABEL_STYLE),
                        rx.text_area(
                            placeholder="Full delivery address...",
                            value=State.f_shipping_address,
                            on_change=State.set_f_shipping_address,
                            rows="3",
                            **{k: v for k, v in INPUT_STYLE.items() if k not in ("padding",)},
                            padding="8px 12px",
                        ),
                        spacing="1",
                        width="100%",
                        align_items="flex-start",
                    ),
                    rx.vstack(
                        labeled_input("No. of Cartons / Boxes *", "e.g. 10", State.f_cartons_count, State.set_f_cartons_count, "number"),
                        labeled_input("Shipping Company *", "e.g. BlueDart, DHL, FedEx", State.f_shipping_company, State.set_f_shipping_company),
                        spacing="3",
                        width="100%",
                    ),
                    spacing="4",
                    width="100%",
                ),

                # Error
                rx.cond(
                    State.form_error != "",
                    rx.text(
                        "⚠ " + State.form_error,
                        color=RED,
                        font_size="13px",
                        font_weight="500",
                    ),
                    rx.box(),
                ),

                # Submit
                rx.hstack(
                    rx.button("💾 Save Entry", on_click=State.submit_entry, **BTN_SUCCESS),
                    rx.button("Cancel", on_click=State.toggle_form, **BTN_OUTLINE),
                    spacing="3",
                ),

                spacing="4",
                width="100%",
                align_items="flex-start",
            ),
            **CARD,
            border=f"1.5px solid {BLUE_LIGHT}",
        ),
        rx.box(),
    )


def filter_bar():
    return rx.box(
        rx.hstack(
            rx.hstack(
                rx.text("📅", font_size="16px"),
                rx.text("From:", font_size="13px", font_weight="600", color=GRAY),
                rx.input(
                    type="date",
                    value=State.filter_date_from,
                    on_change=State.set_filter_date_from,
                    border="1.5px solid #BFDBFE",
                    border_radius="8px",
                    padding="6px 10px",
                    font_size="13px",
                    background="white",
                    _focus={"outline": "none", "border_color": BLUE_MID},
                ),
                rx.text("To:", font_size="13px", font_weight="600", color=GRAY),
                rx.input(
                    type="date",
                    value=State.filter_date_to,
                    on_change=State.set_filter_date_to,
                    border="1.5px solid #BFDBFE",
                    border_radius="8px",
                    padding="6px 10px",
                    font_size="13px",
                    background="white",
                    _focus={"outline": "none", "border_color": BLUE_MID},
                ),
                rx.button("Apply", on_click=State.apply_filter,
                          **{**BTN_PRIMARY, "padding": "7px 16px", "font_size": "13px"}),
                rx.button("Today", on_click=State.reset_to_today,
                          **{**BTN_OUTLINE, "padding": "7px 16px", "font_size": "13px"}),
                spacing="2",
                align_items="center",
                flex_wrap="wrap",
            ),
            rx.spacer(),
            rx.cond(
                State.entries.length() > 0,
                rx.button(
                    "⬇ Export to Excel",
                    on_click=State.export_excel,
                    **{**BTN_SUCCESS, "font_size": "13px", "padding": "7px 18px"},
                ),
                rx.box(),
            ),
            width="100%",
            align_items="center",
            flex_wrap="wrap",
        ),
        **{**CARD, "padding": "16px 24px"},
    )


def summary_bar():
    return rx.cond(
        State.entries.length() > 0,
        rx.hstack(
            summary_card("📋", "Total Entries", State.summary_total),
            summary_card("📦", "Total Cartons", State.summary_cartons),
            rx.box(
                rx.hstack(
                    rx.text("🚚", font_size="28px"),
                    rx.vstack(
                        rx.text(State.summary_companies, font_size="14px", font_weight="600", color=BLUE),
                        rx.text("Shipping Companies", font_size="12px", color=GRAY, font_weight="500"),
                        spacing="0",
                        align_items="flex-start",
                    ),
                    spacing="3",
                    align_items="center",
                ),
                background="white",
                border_radius="12px",
                border=f"1.5px solid {BLUE_LIGHT}",
                padding="16px 22px",
                box_shadow="0 2px 6px rgba(30,64,175,0.07)",
                flex="1",
            ),
            spacing="4",
            width="100%",
            flex_wrap="wrap",
        ),
        rx.box(),
    )


def entries_table():
    return rx.cond(
        State.entries.length() > 0,
        rx.box(
            rx.vstack(
                rx.text(
                    "Shipping Entries",
                    font_size="16px",
                    font_weight="700",
                    color=BLUE,
                    padding_bottom="4px",
                ),
                rx.box(
                    rx.table.root(
                        rx.table.header(
                            rx.table.row(
                                *[
                                    rx.table.column_header_cell(
                                        h,
                                        background=BLUE,
                                        color="white",
                                        font_size="12px",
                                        font_weight="700",
                                        padding="12px 14px",
                                        white_space="nowrap",
                                    )
                                    for h in [
                                        "Name", "Date & Time", "From", "To",
                                        "Shipping Address", "Cartons", "Company", "Action",
                                    ]
                                ]
                            )
                        ),
                        rx.table.body(
                            rx.foreach(State.entries, entry_row)
                        ),
                        width="100%",
                        border_radius="8px",
                        overflow="hidden",
                    ),
                    width="100%",
                    overflow_x="auto",
                ),
                spacing="3",
                width="100%",
                align_items="flex-start",
            ),
            **CARD,
        ),
        rx.box(
            rx.vstack(
                rx.text("📭", font_size="48px"),
                rx.text(
                    "No entries found for the selected date range.",
                    font_size="16px",
                    color=GRAY,
                    font_weight="500",
                ),
                rx.text(
                    "Click '+ Add Entry' to register a new shipment.",
                    font_size="14px",
                    color="#9CA3AF",
                ),
                spacing="2",
                align_items="center",
                padding="40px",
            ),
            **CARD,
            text_align="center",
        ),
    )


# ─────────────────────────────────────────
# PAGE
# ─────────────────────────────────────────

def index() -> rx.Component:
    return rx.box(
        # ── Header ────────────────────────────────────────────────
        rx.box(
            rx.hstack(
                rx.vstack(
                    rx.text(
                        "🏥 Medical Shop",
                        font_size="13px",
                        color="#93C5FD",
                        font_weight="500",
                        letter_spacing="0.05em",
                        text_transform="uppercase",
                    ),
                    rx.text(
                        "Shipping Register",
                        font_size="26px",
                        font_weight="800",
                        color="white",
                        letter_spacing="-0.02em",
                    ),
                    spacing="0",
                    align_items="flex-start",
                ),
                rx.spacer(),
                rx.button(
                    "+ Add Entry",
                    on_click=State.toggle_form,
                    background="white",
                    color=BLUE,
                    border_radius="10px",
                    padding="10px 22px",
                    font_weight="700",
                    font_size="15px",
                    cursor="pointer",
                    border="none",
                    _hover={"background": BLUE_LIGHT},
                ),
                width="100%",
                align_items="center",
            ),
            background=f"linear-gradient(135deg, {BLUE} 0%, #1D4ED8 100%)",
            padding="20px 32px",
        ),

        # ── Body ──────────────────────────────────────────────────
        rx.vstack(
            add_entry_form(),
            filter_bar(),
            summary_bar(),
            entries_table(),
            spacing="4",
            width="100%",
            max_width="1400px",
            margin="0 auto",
            padding="24px 16px 40px",
            align_items="flex-start",
        ),

        background=BG,
        min_height="100vh",
        font_family="'Inter', 'Segoe UI', system-ui, sans-serif",
    )


# ─────────────────────────────────────────
# APP DEFINITION
# ─────────────────────────────────────────

app = rx.App(
    stylesheets=[
        "https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap",
    ],
)
app.add_page(index, route="/", on_load=State.on_load)
